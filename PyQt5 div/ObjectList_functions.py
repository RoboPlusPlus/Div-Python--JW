from openpyxl import workbook, worksheet, load_workbook
import os.path, re
from pathlib import Path
import object_class

def import_rows_from_sheet(path="object_list.xlsx"):
    blank_list = []
    file_exists = os.path.isfile(path)

    if file_exists:
        try:
            row_list = []
            wb = load_workbook(path)
            ws = wb.active
            max_column = ws.max_column
            max_row = ws.max_row
            print("Workbook loaded")
            print("number of columns in imported sheet", max_column)
            print("number of rows in imported sheet", max_row)

            temprow = []
            for row in range(1, max_row+1):
                for column in range(1, max_column+1):
                    cell_value = ws.cell(row=row, column=column).value
                    temprow.append(str(cell_value))

                row_list.append(list(temprow))
                temprow.clear()

            return True, row_list
        except:
            print("failed to import Object list")
            return False, blank_list
    else:
        print("File does not exist")
        return False, blank_list


def clean_sheet_rows(_sheet_rows):
    _removed_rows = []
    for row in reversed(range(len(_sheet_rows))):
        for cell in range(2):
            if "#" in _sheet_rows[row][cell]:
                _removed_rows.append(_sheet_rows[row])
                del _sheet_rows[row]
                break

        for row in _sheet_rows:
            for cell in row:
                cell = str(cell).replace(" ", "")
                cell = str(cell).replace(",", "")
                cell = str(cell).replace("\n", "")
                cell = str(cell).replace("\t", "")

    return _sheet_rows, _removed_rows


def verify_tag_pattern(taglist):
    _verified_tags = []
    _unverified_tags = []
    section_tag_pattern = re.compile("[a-zA-Z][a-zA-Z][a-zA-Z][0-9][0-9]")
    default_tag_pattern = re.compile(
        "[a-zA-Z][a-zA-Z][a-zA-Z][0-9][0-9][a-zA-Z][a-zA-Z][a-zA-Z0-9_][a-zA-Z0-9_][0-9][0-9][0-9][0-9]")
    for entry in taglist:
        tag, object_number = entry
        if section_tag_pattern.match(tag) and len(tag) == 5:
            print("Tag: " + tag + " matched with section tag pattern")
            _input = tag, object_number, "section"
            _verified_tags.append(_input)
            continue
        if default_tag_pattern.match(tag) and len(tag) == 13:
            print("Tag: " + tag + " matched with default tag pattern")
            _input = tag, object_number, "default"
            _verified_tags.append(_input)
            continue
        print("Tag: " + tag + " not matched with any tag pattern")
        _input = tag, object_number, "section"
        _unverified_tags.append(_input)

    return _verified_tags, _unverified_tags


def load_object_io(_path= "C:\\Users\JoachimR\Dropbox\\Programmering Div\Python\\PyQt5 div\\object_setups\object_io_info.xlsx"):
    p = Path(_path)
    loaded_object_types = set()
    blank_list = []
    file_exists = os.path.isfile(_path)

    if file_exists:
    #try:
        row_list = []
        object_objects = []
        wb = load_workbook(_path)
        ws = wb.active
        max_column = ws.max_column
        max_row = ws.max_row
        print("Workbook " + _path + " loaded")
        print("number of columns in imported sheet", max_column)
        print("number of rows in imported sheet", max_row)

        temprow = []
        for row in range(1, max_row+1):
            for column in range(1, max_column+1):
                object_type = ws.cell(row=row, column=1).value
                if object_type not in loaded_object_types and object_type != None and object_type != "objekttype":
                    object_objects.append(object_class.BaseOject(object_type))
                    loaded_object_types.add(object_type)
        for o in object_objects:
            print(o.get_object_type())

        for row in range(1, max_row + 1):
            object_type = ws.cell(row=row, column=1).value
            for i in object_objects:
                if i.get_object_type() == object_type:
                    print("Found " + str(i.get_object_type()) + " == " + str(object_type))
                    didoaiao = ws.cell(row=row, column=2).value
                    function = ws.cell(row=row, column=3).value
                    description = ws.cell(row=row, column=4).value
                    NO_NC_Analog = ws.cell(row=row, column=5).value
                    if str(didoaiao).lower() == "di":
                        i.add_di_entry((didoaiao, function, description, NO_NC_Analog))
                        print("added entry: " + str(didoaiao) + str(function) + str(description) + str(NO_NC_Analog) + " to object_type " + object_type + " Verify: " + i.get_object_type())

        for ob in object_objects:
            print("From object: " + ob.get_object_type())
            print(ob.get_di())


            return True, loaded_object_types, object_objects
        #except:
         #   print("failed to import Object io list")
          #  return False, blank_list
    #else:
     #   print("File does not exist")
        #return False, blank_list, blank_list


def make_db_from_tags():
    pass


def main():
    load_object_io_success, loaded_object_types, object_io_db = load_object_io()
    print(load_object_io_success)
    print(loaded_object_types)
    print(object_io_db)
    """
    import_success, sheet_rows = import_rows_from_sheet()
    print(sheet_rows)
    cleaned_sheet_rows, removed_rows = clean_sheet_rows(sheet_rows)
    print(cleaned_sheet_rows)

    verified_tags, unverified_tags = verify_tag_pattern(cleaned_sheet_rows)
    print(verified_tags)
"""
if __name__ == '__main__':
    main()

