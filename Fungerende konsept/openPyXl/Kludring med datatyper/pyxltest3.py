from openpyxl import Workbook
from openpyxl import load_workbook
import math


 # Variabler
spare_min_prosent = 30
di_card_slots = 8
dq_card_slots = 8
ai_card_slots = 8
ao_card_slots = 8
row_list = []
sheet_list = []
node_names = set()
slave_names = set()
master_names = set()
structs = set()



# Laster xl-ark
wb = load_workbook("Kludrsheet.xlsx")
ws = wb.active
max_column = ws.max_column
max_row = ws.max_row

print("number of columns", max_column)
print("number of rows", max_row)


#laster alle celleverdier inn i "sheet_list". Dette blir en "list of lists" med en liste for hver row.
for i in range (1, max_row +1):
    for j in range(1, max_column + 1):
        row_list.append(ws.cell(row = i, column = j).value)
    sheet_list.append(row_list)
    row_list = []
    

#dekonstruerer sheet_list Deler opp i slave og master.
for k in range(1, max_row):
    _node, _type, _master, _TagName, _IO_type, _Description, _NONC, _Struct, _Subtype, _MB_access = sheet_list[k]
    node_names.add(_node)

    if _master == "IsMaster":
        master_names.add(_node)

    if _master != "IsMaster":
        slave_names.add(_node)



master_names = sorted(master_names)
node_names = sorted(node_names)
slave_names = sorted(slave_names)

dq_list = []
di_list = []
io_list = []

dq_num = 0
di_num = 0
ao_num = 0
ai_num = 0
dq_spare_row = []

for i in range(0, len(master_names)):
    for k in range(1, max_row):
        _node, _type, _master, _TagName, _IO_type, _Description, _NONC, _Struct, _Subtype, _MB_access = sheet_list[k]
        if None ==_Subtype:
            sheet_list[k][8] = ""

        tempmn = master_names[i]
        if tempmn == _node and _IO_type == "DQ":
            dq_list.append(sheet_list[k])
            dq_num = dq_num +1

        if tempmn == _node and _IO_type == "DI":
            di_list.append(sheet_list[k])
            di_num = di_num +1
            
    dq_spare_min = math.ceil(dq_num * (spare_min_prosent/100))
    print(dq_spare_min)
    
    s = 0
    print(len(dq_list))
    while s < dq_spare_min or len(dq_list)%8 != 0 :
        dq_spare_row = [_node, _type, _master, "spare_" + str(len(dq_list)), "DQ", "spare", "NO", "X_List", "", "R"]
        dq_list.append(dq_spare_row)
        s = s+1
    
    print(len(dq_list))


    


    io_list.append(dq_list)
    io_list.append(di_list)
    dq_list = 0
    di_list = 0
    dq_num = 0
    di_num = 0
    ao_num = 0
    ai_num = 0





wb2 = Workbook()

ws2_1= wb2.active

temp_row = 1
for i in range(0, len(master_names)+1):
    for k in range(1, len(io_list[i])+1):
        templist = io_list[i][k-1]
       
        ws2_1.cell(row=temp_row, column =1, value = str(templist[0]))
        ws2_1.cell(row=temp_row, column =2, value = str(templist[1]))
        ws2_1.cell(row=temp_row, column =3, value = str(templist[2]))
        ws2_1.cell(row=temp_row, column =4, value = str(templist[3]))
        ws2_1.cell(row=temp_row, column =5, value = str(templist[4]))
        ws2_1.cell(row=temp_row, column =6, value = str(templist[5]))
        ws2_1.cell(row=temp_row, column =7, value = str(templist[6]))
        ws2_1.cell(row=temp_row, column =8, value = str(templist[7]))
        ws2_1.cell(row=temp_row, column =9, value = str(templist[8]))
        ws2_1.cell(row=temp_row, column =10, value = templist[9])
        temp_row = temp_row + 1


print("lengde IO-list", len(io_list))
print("IO-list 0", len(io_list[0]))
print("IO-list 1", len(io_list[1]))
wb2.save("bla.xlsx")






