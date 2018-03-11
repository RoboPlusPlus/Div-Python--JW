from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import re
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
import math
import string
from datetime import datetime


# Variabler
skap_list = set()
alarm_list = []
minimum_spare_prosent = 15


# Laster xl-ark
wb = load_workbook("1285_001.xlsx")
ws = wb.active
max_column = ws.max_column
max_row = ws.max_row
row_list = []
sheet_list = []
io_total = 0


print("number of columns", max_column)
print("number of rows", max_row)


#laster alle celleverdier inn i "sheet_list". Dette blir en "list of lists" med en liste for hver row.
for i in range (1, max_row +1):
    for j in range(1, max_column + 1):
        row_list.append(ws.cell(row = i, column = j).value)
    sheet_list.append(row_list)
    row_list = []

wb2 = Workbook()
ws2_1= wb2.active

temp_row = 1
ids = []
first_row = True
for row in sheet_list:
    object_name = ""
    prefix = ""
    c1, c2, c3 = row
    address = str(c1)
    identifier = str(c2)
    IDF_Comment = str(c3)
    ids = identifier.split("-")

    object_name_pattern = re.compile("^([A-Z][A-Z][0-9][0-9][0-9]+)+$")
    for i, word in enumerate(ids):
        if len(word.replace(" ", "")) == 5:
            if "." not in word:
                if object_name_pattern.match(word):
                    object_name = word.replace(" ", "")
                    object_position = i

    if len(ids[0]) == 2:
        if not object_name_pattern.match(ids[0]):
            prefix = ids[0]
        else:
            prefix = ""

    if first_row:
        try:
            ws2_1.cell(row=1, column=1, value= "Address") #Kolonne A
            ws2_1.cell(row=1, column=2, value= "identifier") #Kolonne B
            ws2_1.cell(row=1, column=3, value= "object_name") #Kolonne C
            ws2_1.cell(row=1, column=4, value= "tag_prefix") #Kolonne D
            ws2_1.cell(row=1, column=5, value= "tag_postfix_1") #Kolonne E
            ws2_1.cell(row=1, column=6, value= "tag_postfix_2") #Kolonne F
            ws2_1.cell(row=1, column=7, value= "Kommentar") #Kolonne G
            ws2_1.cell(row=1, column=8, value= "5") #Kolonne H
            ws2_1.cell(row=1, column=9, value= "6") #Kolonne I
            ws2_1.cell(row=1, column=10, value= "7") #Kolonne J
            #first_row = False

        except:
            pass
    try:
        if address != "None":
            ws2_1.cell(row=temp_row, column=1, value=str(address))
    except:
        pass
    try:
        if identifier != "None":
            ws2_1.cell(row=temp_row, column=2, value=str(identifier))
    except:
        pass
    try:
        if object_name != "None":
            ws2_1.cell(row=temp_row, column=3, value=str(object_name))
    except:
        pass
    try:
        if ids[0] != "None":
            if object_position > 0:
                if prefix != "":
                    ws2_1.cell(row=temp_row, column=4, value=str(ids[0].replace(".", "")))
    except:
        pass
    try:
        if object_position < 2:
            if ids[1] != "None":
                if ids[1] != object_name:
                    ws2_1.cell(row=temp_row, column=5, value=str(ids[1].replace(".", "")))
    except:
        pass
    try:
        if ids[2] != "None":
            ws2_1.cell(row=temp_row, column=6, value=str(ids[2].replace(".", "")))
    except:
        pass
    try:
        if ids[3] != "None":
            ws2_1.cell(row=temp_row, column=7, value=str(ids[3].replace(".", "")))
    except:
        pass
    try:
        if ids[4] != "None":
            ws2_1.cell(row=temp_row, column=8, value=str(ids[4].replace(".", "")))
    except:
        pass
    try:
        if IDF_Comment != "None":
            ws2_1.cell(row=temp_row, column=7, value=str(IDF_Comment.replace("]", "AA")))

    except:
        pass
    ids.clear()
    temp_row +=1

# define a table style
mediumStyle =TableStyleInfo(name='TableStyleMedium2',
                                                      showRowStripes=True)
# create a table
table = Table(ref='A1:J520',
              displayName='FruitColors',
              tableStyleInfo=mediumStyle)

# add the table to the worksheet
ws2_1.add_table(table)


for col in ws2_1.columns:
    max_length = 0
    column = col[0].column # Get the column name
    for cell in col:
        if cell.coordinate in ws2_1.merged_cells: # not check merge_cells
            continue
        try: # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws2_1.column_dimensions[column].width = adjusted_width

wb2.save("sjekk.xlsx")



