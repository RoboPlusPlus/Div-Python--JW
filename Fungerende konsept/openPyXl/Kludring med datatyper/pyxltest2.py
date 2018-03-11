from openpyxl import Workbook
from openpyxl import load_workbook


 # Variabler
row_list = []
sheet_list = []
node_names = set()
slave_names = set()
master_names = set()
structs = set()



# Laster xl-ark
wb = load_workbook("Kludrsheet.xlsx")
ws = wb.active
max_column = ws.max_column
max_row = ws.max_row

print("number of columns", max_column)
print("number of rows", max_row)


#laster alle celleverdier inn i "sheet_list". Dette blir en "list of lists" med en liste for hver row.
for i in range (1, max_row +1):
    for j in range(1, max_column + 1):
        row_list.append(ws.cell(row = i, column = j).value)
    sheet_list.append(row_list)
    row_list = []
    

#dekonstruerer sheet_list Deler opp i slave og master.
for k in range(1, max_row):
    _node, _type, _master, _TagName, _IO_type, _Description, _NONC, _Struct, _Subtype, _MB_access = sheet_list[k]
    node_names.add(_node)

    if _master == "IsMaster":
        master_names.add(_node)

    if _master != "IsMaster":
        slave_names.add(_node)



master_names = sorted(master_names)
node_names = sorted(node_names)
slave_names = sorted(slave_names)

dq_list = []
di_list = []
io_list = []

for i in range(0, len(master_names)):
    for k in range(1, max_row):
        _node, _type, _master, _TagName, _IO_type, _Description, _NONC, _Struct, _Subtype, _MB_access = sheet_list[k]

        tempmn = master_names[i]
        if tempmn == _node and _IO_type == "DQ":
            dq_list.append(sheet_list[k])

        if tempmn == _node and _IO_type == "DI":
            di_list.append(sheet_list[k])

    io_list.append(dq_list)
    io_list.append(di_list)
    dq_list = 0
    di_list = 0

print(io_list)



wb2 = Workbook()

ws2_1= wb2.active

ws2_1.cell

wb2.save("bla.xlsx")
