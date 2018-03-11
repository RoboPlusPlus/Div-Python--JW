from openpyxl import Workbook
from openpyxl import load_workbook
import math
import string
from datetime import datetime


# Variabler
skap_list = set()
alarm_list = []
minimum_spare_prosent = 15


# Laster xl-ark
wb = load_workbook("New Microsoft Excel Worksheet.xlsx")
ws = wb.active
max_column = ws.max_column
max_row = ws.max_row
row_list = []
sheet_list = []
io_total = 0


print("number of columns", max_column)
print("number of rows", max_row)


#laster alle celleverdier inn i "sheet_list". Dette blir en "list of lists" med en liste for hver row.
for i in range (1, max_row +1):
    for j in range(1, max_column + 1):
        row_list.append(ws.cell(row = i, column = j).value)
    sheet_list.append(row_list)
    row_list = []




    
siste_sjekkede_kanal_nummer = 0
siste_sjekkede_kort_nummer = 0
#dekonstruerer sheet_list, og finner alle fordelinger
for k in range(3, max_row):
    
    skaptype, skap_nr, Prefiks, Avsnitt, Type, Subtype, Enhetsnummer, \
    Sett_OPC_tag_Manuelt, Kameranummer, Kamerafrys, Funksjon, OPC_tag, \
    signalbeskrivelse, IO, signaltype, måleområde_min, måleområde_max, \
    måleenhet, modul_nr, kort_nr, kanal_nr, Type_kort, endring_i_forhold_til_tilbud, \
    rev  = sheet_list[k]
    
    
    io_total = io_total +1
    beskrivelse = signalbeskrivelse
    IO_type =  IO
    fordeling = str(skaptype) + str(skap_nr)
    #print(beskrivelse)
    skap_list.add(fordeling)

    if skaptype == None:
        error_to_report = ("Skaptype mangler på rad " + str(k + 1))
        alarm_list.append(error_to_report)

    if skap_nr == None:
        error_to_report = ("Skapnummer mangler på rad " + str(k + 1))
        alarm_list.append(error_to_report)
        

    if Funksjon == None:
        error_to_report = ("Funksjon mangler på rad " + str(k + 1))
        alarm_list.append(error_to_report)

    if OPC_tag == "":
        if str(Funksjon).lower() != "spare":
            error_to_report = ("OPC-tag mangler på rad " + str(k + 1))
            alarm_list.append(error_to_report)
    
    #if IO == None:
     #   error_to_report = ("IO-type mangler på rad " + str(k + 1))
      #  alarm_list.append(error_to_report)

    if modul_nr == None:
        error_to_report = ("Modul_nr mangler på rad " + str(k + 1))
        alarm_list.append(error_to_report)


    if kort_nr == None:
        error_to_report = ("Kort-nummer mangler på rad " + str(k + 1))
        alarm_list.append(error_to_report)        


    if kanal_nr == None:
        error_to_report = ("Kanal-nummer mangler på rad " + str(k + 1))
        alarm_list.append(error_to_report)

    if kanal_nr != None:
        if k > 4:
            if int(kanal_nr) == 0 :
                if str(siste_sjekkede_kanal_nummer) != "1":
                    if (((siste_sjekkede_kanal_nummer +1)%4 != 0.0)):         
                        error_to_report = ("Rart kanal-nummer på rad " + str(k + 1))
                        alarm_list.append(error_to_report)    
                    
    if kanal_nr != None:
        if siste_sjekkede_kanal_nummer != None:
            if siste_sjekkede_kanal_nummer >= kanal_nr & kanal_nr != 0:
                error_to_report = ("Kanal-nummer med verdi: " + str(kanal_nr) + " på rad " + str(k) + " virker feil")
                alarm_list.append(error_to_report)
                
    if kanal_nr != None:
        if siste_sjekkede_kanal_nummer != None:
            if kort_nr != siste_sjekkede_kort_nummer:
                if str(kanal_nr) != "0":
                    error_to_report = ("Kanal eller kort-nummer på rad " + str(k+1) + " virker feil")
                    alarm_list.append(error_to_report)
                    
    if kanal_nr != None:
        if siste_sjekkede_kanal_nummer != None:
            if kort_nr > siste_sjekkede_kort_nummer + 1:
                    error_to_report = ("Kanal eller kort-nummer på rad " + str(k+1) + " virker feil")
                    alarm_list.append(error_to_report)    


    

    siste_sjekkede_kanal_nummer = kanal_nr
    siste_sjekkede_kort_nummer = kort_nr
        
skap_list = sorted(skap_list)

spare_total = 0

print("\n")
for skap in skap_list:
    DI_count = 0
    DO_count = 0
    DI_Spare_count = 0
    DO_Spare_count = 0
    di_spare_prosent = 0.0
    do_spare_prosent = 0.0
        
    print("#######\n"+skap+"\n#######")
    for k in range(3, max_row):
       

        #dekomponering av tuple som representerer rad
        skaptype, skap_nr, Prefiks, Avsnitt, Type, Subtype, Enhetsnummer, Sett_OPC_tag_Manuelt, Kameranummer, Kamerafrys, funksjon, OPC_tag, signalbeskrivelse, IO, signaltype, måleområde_min, måleområde_max, måleenhet, modul_nr, kort_nr, kanal_nr, Type_kort, endring_i_forhold_til_tilbud, rev  = sheet_list[k]

        beskrivelse = funksjon
        IO_type =  IO
        fordeling = str(skaptype) + str(skap_nr)


    #################################
    #
    #Beregning av IO-antall og spare
    #
    #################################
        if skap == fordeling:
            if str(IO_type).lower() == "di":
                DI_count = DI_count + 1
                if str(beskrivelse).lower() == "spare":
                    DI_Spare_count = DI_Spare_count +1
                    spare_total = spare_total +1
            if str(IO_type).lower() == "do":
                DO_count = DO_count +1
                if str(beskrivelse).lower() == "spare":
                    DO_Spare_count = DO_Spare_count +1
                    spare_total = spare_total +1

                    
    print(skap + " har " + str(DI_count) + "DI")
    print(skap + " har " + str(DO_count) + "DO")
    print(skap + " har " + str(DI_Spare_count) + " DI spare")
    print(skap + " har " + str(DO_Spare_count) + " DO spare")




    string.Formatter.convert_field.__subclasshook__
    if  DI_count> 0:
        di_spare_prosent= (DI_Spare_count / DI_count)*100

        if di_spare_prosent < minimum_spare_prosent:
            
            error_to_report = str(skap + " har " + str(di_spare_prosent) + " prosent DI-spare")
            alarm_list.append(error_to_report)

        
        

    if  DO_count> 0:
        do_spare_prosent= (DO_Spare_count / DO_count)*100

        if do_spare_prosent < minimum_spare_prosent:
            
            error_to_report = str(skap + " har " + str(do_spare_prosent) + " prosent DO-spare")
            alarm_list.append(error_to_report)

    
    
    print(skap + " har " + str(di_spare_prosent) + "prosent DI-spare")
    print(skap + " har " + str(do_spare_prosent) + "prosent DO-spare")
    print("\n")

    

print("total number of IO: " + str(io_total))
print("total number of spare: " + str(spare_total))
print("total spare prosent: " + str((spare_total / io_total)*100))



#year_now, months_now, days_now, hours_now, minutes_now, seconds_now, ms_now =
time_now = datetime.minute()



print("#######\nAlarmlogg\n#######")
for alarm in alarm_list:
    print(alarm)

print(time_now)







