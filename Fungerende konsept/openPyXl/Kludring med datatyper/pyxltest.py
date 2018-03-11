from openpyxl import Workbook
from openpyxl import load_workbook


wb = load_workbook("Kludrsheet.xlsx")
ws = wb.active

max_column = ws.max_column
max_row = ws.max_row


a = ws.values

print ("number of columns", max_column)
print ("number of rows", max_row)
row_list= []
sheet_list = []
node_names = set()
master_names = set()
structs = set()
di_list = []
dq_list = []
templist = []
masterDQlist = []


       
for i in range (1, max_row +1):
    for j in range(1, max_column +1):
        row_list.append(ws.cell(row = i, column = j).value)
    sheet_list.append(row_list)
    row_list = []
    


for k in range (1, max_row):
    _node, _type, _master, _TagName, _IO_type, _Description, _NONC, _Struct, _Subtype, _MB_access = sheet_list[k]
    node_names.add(_node)
    if _IO_type == "DI":
        di_list.append(sheet_list[k])
    if _IO_type == "DQ":
        dq_list.append(sheet_list[k])
    if _master == "IsMaster":
        master_names.add(_node)
        if _IO_type == "DQ":
            masterDQlist.append(sheet_list[k])
        

master_names = sorted(master_names)

for i in range(0, len(master_names)):
    for j in range (1, max_row):
        _node, _type, _master, _TagName, _IO_type, _Description, _NONC, _Struct, _Subtype, _MB_access = sheet_list[j]
        if master_names[i] == _node:
            print(_node)
            


    
#print(masterDQlist)
wb2 = Workbook()

ws2_1= wb2.active

wb2.save("bla.xlsx")
