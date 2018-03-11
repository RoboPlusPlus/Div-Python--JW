from openpyxl import load_workbook
from openpyxl import Workbook
import datetime
wb = load_workbook('hest.xlsx') #Henter inn workbook
ws =wb.active   #sier at ws= aktiv workbook. Bør/skal være mer presist
ws['A1'] = 49
ws['A2'] = '=A1+1'



wbOut = Workbook()



#Kode for å legge til worksheets, samt å sjekke at dette sheetet ikke finnes fremdeles.
#Hvis ws finnes, så slett.

sheetnames = ("IO_sort", "Structs", "MB_Adresses_IO")

print(wb.sheetnames)



#ws_calc = wb.create_sheet(title="calculations")
ws_ioSort = wb.create_sheet(title=sheetnames[0])

a= len(sheetnames)
ws_ioSort['A1'] = a

now = datetime.datetime.now()
print (now.year)
print ( "Number of rows on ws " + str(ws.max_row))


wsO1 = wbOut.active
#end = wb.copy_worksheet("Sheet")
#wsO1 = end
name = "wbOut" + str(now.year)+"_" + str(now.month)+"_" + str(now.day)+"_" + str(now.hour)+"_" + str(now.minute) + ".xlsx"
wbSaveName = (name)
print (wbSaveName)
wbOut.save(wbSaveName)
