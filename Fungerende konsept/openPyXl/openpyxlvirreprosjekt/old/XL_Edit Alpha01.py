from openpyxl import load_workbook
from openpyxl import Workbook
import numpy as np
import datetime

#starter med å importere workbook hest inn i minnnet
wb = load_workbook('hest.xlsx') #Henter inn workbook


#Finner sheet

#analyserer rader og kolonner i alle sheets
for x in range (0, len(wb.worksheets)):
    sheet = wb.worksheets[x]

    row_count = sheet.max_row
    column_count = sheet.max_column
    
    print (wb.sheetnames[0])
    print ("Number of rows: " + str(row_count))
    print ("Number of columns: " + str(column_count) +"\n")

#Lagrer celler fra sheets i en tuple

sheetDump =np.zeros((50,500,500))
numOfSheets = len(wb.worksheets)
sheetnumber = 0
x= 0
for x in range (0, numOfSheets):
    sheet = wb.worksheets[x]
    sheetnumber = sheetnumber +1
    x= x+1
    for y in range (0, sheet.max_row):
        y= y+1
        for z in range (0, sheet.max_column):
            z=z+1
            #sheetDump[(x,y,z)] = sheet.cell(row=y, column=z)"worksheet " + str(sheetnumber) + 
            strtemp = sheet.cell(row=y, column=z)
            print(str(strtemp))


