from openpyxl import Workbook
wb = Workbook()
ws =wb.open("hest.xlsx")

ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])
ws.append([4, 5, "kuli"])
# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")

ws1 = wb.create_sheet("Testsheet") # insert at the end (default)

print(wb.sheetnames)


wb.save("hest.xlsx")
