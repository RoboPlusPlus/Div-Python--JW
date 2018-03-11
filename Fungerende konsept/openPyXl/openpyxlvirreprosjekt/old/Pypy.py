from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl.utils
import numpy as np
import datetime

#starter med å importere workbook hest inn i minnnet
wb = load_workbook('hest.xlsx') #Henter inn workbook


#Finner sheet

#analyserer rader og kolonner i alle sheets
for x in range (0, len(wb.worksheets)):
    sheet = wb.worksheets[x]
    row_count = sheet.max_row
    column_count = sheet.max_column

    
    print (wb.sheetnames[x])
    print ("Number of rows: " + str(row_count))
    print ("Number of columns: " + str(column_count) +"\n")



ws = wb.worksheets[3]

print(ws.max_row)
x=0
for x in range (1, ws.max_row):
    x= x+1
    for y in range (1, ws.max_column):
        d = ws.cell(row = x, column = y)
        y=y+1
        print (d.value)
