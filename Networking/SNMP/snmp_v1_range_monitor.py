from openpyxl import Workbook
from openpyxl import load_workbook
import subprocess
import socket


sheet_list = []
row_list = []
# Laster xl-ark
wb = load_workbook("Snmp_v1_monitor_list.xlsx")
ws = wb.active
max_column = ws.max_column
max_row = ws.max_row
ip_list = []
scan_setup_list = []

print("number of columns", max_column)
print("number of rows", max_row)


#laster alle celleverdier inn i "sheet_list". Dette blir en "list of lists" med en liste for hver row.
for i in range (1, max_row +1):
    for j in range(1, max_column + 1):
        row_list.append(ws.cell(row = i, column = j).value)
    sheet_list.append(row_list)

for i in range (1, max_row +1):
        ip_list.append(ws.cell(row = i, column = 1).value)

for i in range (1, max_row +1):
    scan_setup_list.append(ws.cell(row = i, column = 2).value)


print(ip_list)
print(scan_setup_list)

#++++++
#SNMPv1
#++++++

from pysnmp.hlapi import *

#Samling av forespørsler.
MIB_refs = ["sysDescr"]#, "sysObjectID","sysUpTime","sysContact","sysName","sysLocation","sysServices", "sysORLastChange", "snmpInPkts", "snmpBasicCompliance"]
#for ip in ip_list:

def snmp_scan(_ip):
    for MIB in MIB_refs:
        errorIndication, errorStatus, errorIndex, varBinds = next(
            getCmd(SnmpEngine(),
                   CommunityData('public', mpModel=0),
                   UdpTransportTarget((_ip, 161)),
                   ContextData(),

                   ObjectType(ObjectIdentity('SNMPv2-MIB', MIB, 0)))
        )

        if errorIndication:
            print(errorIndication)
        elif errorStatus:
            print('%s at %s' % (errorStatus.prettyPrint(),
                                errorIndex and varBinds[int(errorIndex) - 1][0] or '?'))
        else:
            for varBind in varBinds:
                print(' = '.join([x.prettyPrint() for x in varBind]))



import threading, time, pprint, os.path, subprocess, sys
from time import localtime, strftime

print(strftime("%a, %d %b %Y %H:%M:%S,", localtime()))  # 'Thu, 28 Jun 2001 14:17:15'


def pinger(_host, _number_of_pings="1"):
    """
    :param _host: IP or Hostname
    :param _number_of_pings: Number of pings before returning
    :return: _ping_answer_from_host(bool), _ping_statistics(dict), _host(string)

    """

    _ping_answer_from_host = False
    _replies = []

    _ping_statistics = {
        "Packets_sent": -1,
        "Received": -1,
        "Lost": -1,
        "Round trip Minimum": -1,
        "Round trip Maximum": -1,
        "Round trip Average": -1,
        "Error": ""
    }

    try:
        _ping = subprocess.Popen(["ping", "-n", _number_of_pings, _host], stdout=subprocess.PIPE,
                                 stderr=subprocess.PIPE, shell=True)
        out, error = _ping.communicate()
        out = out.strip()
        error = error.strip()
        _ping_statistics["Error"] = error.decode('utf-8')
        subprocess_success = True
    except:
        _ping_statistics["Error"] = "Could not successfully run subprocess"
        raise Exception("Could not successfully run subprocess")

    if subprocess_success:
        # Analyserer bytestring med ping-resultat
        _ping_result = out.decode('utf-8')
        _ping_result_lines = _ping_result.split("\n")

        for line in _ping_result_lines:
            if line.startswith("Reply from "):
                _replies.append(line)

            if line.find("Destination host unreachable."):
                _replies.append(line)

            if line.startswith("    Packets: Sent = "):
                _fragmented_line = line.split(" ")
                _ping_statistics["Packets_sent"] = _fragmented_line[7].replace(",", "")
                _ping_statistics["Received"] = _fragmented_line[10].replace(",", "")
                _ping_statistics["Lost"] = _fragmented_line[13]

            if line.startswith("    Minimum = "):
                _fragmented_roundtrip_line = line.split(" ")
                _ping_statistics["Round trip Minimum"] = _fragmented_roundtrip_line[6].replace("ms", "").replace(",",
                                                                                                                 "")
                _ping_statistics["Round trip Maximum"] = _fragmented_roundtrip_line[9].replace("ms", "")
                _ping_statistics["Round trip Average"] = _fragmented_roundtrip_line[12].replace("ms", "")

            if int(_ping_statistics["Round trip Minimum"]) >= 0:
                _ping_answer_from_host = True

    return _ping_answer_from_host, _ping_statistics, _host




access_pingstatus_lock = threading.Lock()
threads = []
ping_statistics = {}

def check_port_161(_ip):
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    result = sock.connect_ex((_ip, 161))
    if result == 0:
        print("Port is open")
    else:
        print("Port is not open")

def main():
    for ip in ip_list:
        _xsuccess, _xping_statistics, _xhost = pinger(ip)
        if not _xsuccess:
            print("No response from: " + ip)

        if _xsuccess:
            print("Ping-response from: " + _xhost + " with statistics: ")
            print(_xping_statistics)



            #print("starting snmp-scan\n")
            #snmp_scan(ip)

#main()
check_port_161('10.181.5.29')